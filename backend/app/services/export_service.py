import io
import json
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class ExportService:
    @staticmethod
    def generate_xlsx(results: list) -> bytes:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ranking Results"
        
        # Headers
        headers = [
            "Rank", "Candidate Name", "Current Role", "Current Company",
            "Experience", "Location", "Match %", "AI Score", "Eligibility",
            "Critical Skill Coverage", "Top Skills", "Recommendation", "Reasoning"
        ]
        ws.append(headers)
        
        # Styling headers
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            
        # Data rows
        for row in results:
            cand_name = row.get("candidate_name") or row.get("candidate_id", "")
            role = row.get("current_title", "")
            company = row.get("current_company", "")
            exp = f"{row.get('years_of_experience', 0.0)} Years"
            loc = row.get("location", "")
            match_pct = f"{row.get('match_percent')}%"
            ai_score = row.get("ai_score", 0.0)
            elig = "Eligible" if row.get("eligibility") else f"Ineligible: {row.get('eligibility_reason', '')}"
            skill_cov = row.get("critical_skill_coverage", "")
            
            skills_raw = row.get("top_skills", [])
            skills_list = [s.get("name") if isinstance(s, dict) else str(s) for s in skills_raw]
            skills_str = ", ".join(skills_list)
            
            recommendation = row.get("hiring_readiness", "")
            reasoning = row.get("reasoning", "")
            
            ws.append([
                row.get("rank"), cand_name, role, company,
                exp, loc, match_pct, ai_score, elig,
                skill_cov, skills_str, recommendation, reasoning
            ])
            
        # Auto-adjust column width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                # Handle cell value being a float or numeric in length check
                max_len = max(max_len, len(val))
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)
            
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def generate_pdf(project_name: str, job_title: str, results: list) -> bytes:
        buffer = io.BytesIO()
        # Landscape might be better for wide layout, but letter portrait is standard.
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        story = []
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'DocTitle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=24,
            textColor=colors.HexColor('#1A365D'),
            spaceAfter=12
        )
        subtitle_style = ParagraphStyle(
            'DocSubTitle',
            parent=styles['Normal'],
            fontName='Helvetica-Oblique',
            fontSize=12,
            textColor=colors.HexColor('#4A5568'),
            spaceAfter=24
        )
        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=colors.HexColor('#2B6CB0'),
            spaceBefore=12,
            spaceAfter=8
        )
        body_style = ParagraphStyle(
            'Body',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=10,
            textColor=colors.HexColor('#2D3748'),
            spaceAfter=6
        )
        table_header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=9,
            textColor=colors.white
        )
        table_cell_style = ParagraphStyle(
            'TableCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=8,
            textColor=colors.HexColor('#2D3748')
        )
        
        story.append(Paragraph("HireMind AI Recruiter Summary", title_style))
        story.append(Paragraph(f"Project: {project_name} | Job Description: {job_title}", subtitle_style))
        story.append(Spacer(1, 12))
        
        # Render executive summary
        story.append(Paragraph("Executive Summary", section_heading))
        story.append(Paragraph(f"Total evaluated candidates: {len(results)}. Below are the top ranked matches based on semantic fit, experience alignment, and critical skill coverage.", body_style))
        story.append(Spacer(1, 12))
        
        # Table of top matches
        story.append(Paragraph("Top Candidates Ranking", section_heading))
        
        # Table headers
        data = [[
            Paragraph("Rank", table_header_style),
            Paragraph("Name", table_header_style),
            Paragraph("Current Role", table_header_style),
            Paragraph("Exp", table_header_style),
            Paragraph("Match", table_header_style),
            Paragraph("Eligibility", table_header_style),
            Paragraph("Readiness", table_header_style)
        ]]
        
        for row in results[:20]: # Limit to top 20 in PDF
            cand_name = row.get("candidate_name") or row.get("candidate_id", "")
            role = row.get("current_title", "")
            exp = f"{row.get('years_of_experience', 0.0)} yrs"
            match_pct = f"{row.get('match_percent')}%"
            elig = "Eligible" if row.get("eligibility") else "Ineligible"
            readiness = row.get("hiring_readiness", "")
            
            data.append([
                Paragraph(f"#{row.get('rank')}", table_cell_style),
                Paragraph(cand_name, table_cell_style),
                Paragraph(role[:25], table_cell_style),
                Paragraph(exp, table_cell_style),
                Paragraph(match_pct, table_cell_style),
                Paragraph(elig, table_cell_style),
                Paragraph(readiness, table_cell_style)
            ])
            
        t = Table(data, colWidths=[40, 100, 130, 45, 45, 60, 60])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2B6CB0')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F7FAFC')]),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        story.append(t)
        
        # Detail analysis page
        story.append(PageBreak())
        story.append(Paragraph("Detailed Top Match Analysis", section_heading))
        story.append(Spacer(1, 8))
        
        for row in results[:3]:
            cand_name = row.get("candidate_name") or row.get("candidate_id", "")
            story.append(Paragraph(f"<b>Rank #{row.get('rank')} - {cand_name}</b> (Score: {row.get('match_percent')}% | {row.get('hiring_readiness', 'medium')} readiness)", body_style))
            story.append(Paragraph(f"<b>Reasoning:</b> {row.get('reasoning', '')}", body_style))
            
            strengths = row.get("strengths", [])
            if strengths:
                story.append(Paragraph(f"• <b>Key Strengths:</b> {', '.join(strengths)}", body_style))
            weaknesses = row.get("weaknesses", [])
            if weaknesses:
                story.append(Paragraph(f"• <b>Areas of Concern:</b> {', '.join(weaknesses)}", body_style))
                
            story.append(Spacer(1, 10))
            
        doc.build(story)
        return buffer.getvalue()
